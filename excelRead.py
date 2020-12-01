#!/usr/bin/env python
# encoding: utf-8
'''
@author: yanghong
@file: excelRead.py
@time: 2020/6/16 13:47
@desc:
'''
from openpyxl import load_workbook


class Excel_read:
    def __init__(self, filepath, sheetname):
        self.wk = load_workbook(filepath)
        self.sheet = self.wk[sheetname]
        self.maxrow = self.sheet.max_row  # 最大行数
        self.maxcolumn = self.sheet.max_column  # 最大列数

    def getAllData(self):
        if self.maxrow <= 1:
            print('总行数小于1')
        else:
            allData = []
            for i in range(2, self.maxrow + 1):
                lineData = []
                for j in range(1, self.maxcolumn + 1):
                    value = self.sheet.cell(i, j).value
                    lineData.append(value)
                allData.append(lineData)
            return (allData)


if __name__ == "__main__":
    a = Excel_read('logindata.xlsx', 'Sheet1')
    print(a.getAllData())
