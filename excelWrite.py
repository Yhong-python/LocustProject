#!/usr/bin/env python
# encoding: utf-8
'''
@author: yanghong
@file: excelWrite.py
@time: 2020/12/1 14:55
@desc:
'''
# 多线程写入数据到excel

import time
from multiprocessing import Pool
from openpyxl import load_workbook

class Write_excel:
    def __init__(self, excelpath):
        self.excelpath = excelpath
        self.wb = load_workbook(self.excelpath)
        self.ws = self.wb.active

    def write(self, row, col, value):
        try:
            self.ws.cell(row, col).value = value
            # self.wb.save(self.excelpath)
        except:
            self.close()
            # self.wb.close()

    def close(self):
        self.wb.close()
    def save(self):
        self.wb.save(self.excelpath)

def mycallback(line, username, i):
    excel.write(line, 1, username)
    excel.write(line, 2, i)

def sayHi(line, username, i):
    print(line, username, i)
    return line, username, i

if __name__ == '__main__':
    pool = Pool(4)
    excel = Write_excel('234.xlsx')
    line = 2  # 从第二行开始写入数据
    time1 = time.time()
    for i in range(100007, 106008):
        username = '18300' + str(i)
        #多进程写入文件时一定要使用callback回调函数，不然进程间会插入不了数据
        pool.apply_async(sayHi, (line, username, i), callback=mycallback(line, username, i))
        line += 1
    pool.close()
    pool.join()
    excel.save()
    excel.close()
    print(time.time() - time1)
